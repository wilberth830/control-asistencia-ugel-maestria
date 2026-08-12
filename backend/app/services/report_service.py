"""TEC-D09 + TEC-D12 — annex reports (replicación exacta de formato UGEL San Román)."""

from __future__ import annotations

from calendar import monthrange
from collections import Counter, defaultdict
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.repositories.institution_repository import institution_repository
from app.repositories.oracle import OracleRepositoryError
from app.services.attendance_service import attendance_service
from app.services.staff_member_service import (
    StaffMemberNotFoundError,
    staff_member_service,
)

DEMO_INSTITUTION = {
    "ugel": "SAN ROMAN",
    "school_name": "Nombre de IE",
    "modular_code": "Codigo modular",
    "education_level": "Nivel / Modalidad",
    "shift_name": "Turno_of",
    "address": "Direccion de IE",
    "department": "PUNO",
    "province": "SAN ROMAN",
    "district": "",
}

MONTH_NAMES = (
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
)

LEGAL_HEADER = (
    "NORMAS PARA EL REGISTRO Y CONTROL DE ASISTENCIA Y SU APLICACION EN LA "
    "PLANILLA UNICA DE PAGOS DE LOS PROFESORES Y AUXILIARES DE EDUCACION, "
    "EN EL MARCO DE LA LEY DE REFORMA MAGISTERIAL Y SU REGLAMENTO "
    "(R.S.G. N 326-2017-MINEDU)"
)

FILL_TITLE = PatternFill("solid", fgColor="008BCE")
FILL_HEADER = PatternFill("solid", fgColor="BDD7EE")
FILL_SUB = PatternFill("solid", fgColor="D6EAF8")
FILL_WEEKEND = PatternFill("solid", fgColor="FCE4D6")
FONT_BLACK_BOLD = Font(bold=True, size=8, name="Calibri")
FONT_RED_BOLD = Font(bold=True, color="C00000", size=9, name="Calibri")
FONT_RED_WEEKEND = Font(bold=True, color="C00000", size=7.5, name="Calibri")
FONT_TITLE = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
FONT_NORMAL = Font(size=8, name="Calibri")
FONT_TINY = Font(size=7.5, name="Calibri")
THIN_BLACK = Side(style="thin", color="000000")
BORDER_THIN = Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
NOWRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _safe(value: Any, default: str = "") -> str:
    if value is None:
        return default
    text = str(value)
    replacements = {
        "Ã±": "n", "Ã‘": "N", "Ã¡": "a", "Ã©": "e", "Ã­": "i", "Ã³": "o", "Ãº": "u",
        "ñ": "n", "Ñ": "N", "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def _segments(end_col: int, n: int) -> list[tuple[int, int]]:
    """Divide 1..end_col en n segmentos proporcionales (el ultimo absorbe el resto)."""
    base = end_col // n
    segs = []
    start = 1
    for i in range(n):
        width = base if i < n - 1 else end_col - start + 1
        segs.append((start, start + width - 1))
        start += width
    return segs


def _place_pair(
    sheet: Any,
    row: int,
    start_col: int,
    end_col: int,
    label: str,
    value: str,
    label_width: int = 2,
) -> None:
    """Coloca un par 'ETIQUETA: valor' dentro de [start_col, end_col], fusionando ambas partes."""
    label_end = min(start_col + label_width - 1, max(start_col, end_col - 1))
    if label_end > start_col:
        sheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=label_end)
    c = sheet.cell(row, start_col, label)
    c.font = FONT_BLACK_BOLD
    c.alignment = NOWRAP_LEFT
    value_start = label_end + 1
    if value_start <= end_col:
        if end_col > value_start:
            sheet.merge_cells(start_row=row, start_column=value_start, end_row=row, end_column=end_col)
        v = sheet.cell(row, value_start, value)
        v.font = FONT_RED_BOLD
        v.alignment = NOWRAP_LEFT


class ReportService:
    def annex_03(
        self,
        month: int,
        year: int,
        institution: dict[str, Any] | None = None,
        import_id: int | None = None,
    ) -> dict[str, Any]:
        inst = institution or self._institution()
        attendance_rows = attendance_service.list_month(month, year, import_id=import_id)
        days_by_staff: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for row in attendance_rows:
            days_by_staff[row["staff_member_id"]].append(row)
        rows = []
        # Todo personal activo (inactivos fuera). Sin marcaciones => days vacio.
        for staff in staff_member_service.list(is_active="Y"):
            days = days_by_staff.get(staff["id"], [])
            rows.append(
                {
                    "staff_member_id": staff["id"],
                    "dni": staff.get("dni"),
                    "full_name": self._full_name(staff),
                    "job_title": staff.get("job_title"),
                    "employment_status": staff.get("employment_status"),
                    "days": sorted(days, key=lambda day: day["attendance_date"]),
                }
            )
        return {
            "institution": inst,
            "period": {"month": month, "year": year},
            "source": "attendance_day",
            "rows": rows,
        }

    def annex_04(self, month: int, year: int, import_id: int | None = None) -> dict[str, Any]:
        attendance_rows = attendance_service.list_month(month, year, import_id=import_id)
        totals = Counter(row["status"] for row in attendance_rows)
        staff_member_ids = {row["staff_member_id"] for row in attendance_rows}
        return {
            "institution": self._institution(),
            "period": {"month": month, "year": year},
            "source": "attendance_day",
            "staff_count": len(staff_member_ids),
            "totals": {
                "present": totals["present"],
                "late": totals["late"],
                "absent": totals["absent"],
                "justified": totals["justified"],
                "leave": totals["leave"],
                "unpaid_leave": totals["unpaid_leave"],
                "permission": totals["permission"],
                "strike": totals["strike"],
                "holiday": totals["holiday"],
                "no_record": totals["no_record"],
            },
        }

    def monthly_workbook(
        self,
        month: int,
        year: int,
        import_id: int | None = None,
        institution_override: dict[str, Any] | None = None,
    ) -> BytesIO:
        if month < 1 or month > 12:
            raise ValueError("invalid_month")
        institution = {**self._institution()}
        if institution_override:
            for key in (
                "ugel", "school_name", "modular_code", "education_level",
                "shift_name", "address", "department", "province", "district",
            ):
                if key in institution_override and institution_override[key] is not None:
                    institution[key] = institution_override[key]
        staff_rows = self._staff_rows(month, year, import_id=import_id)
        workbook = Workbook()
        sh3 = workbook.active
        sh3.title = "ASISTENCIA"
        sh4 = workbook.create_sheet("REPORTE CONSOLIDADO")
        self._write_attendance_sheet(sh3, institution, staff_rows, month, year)
        self._write_consolidated_sheet(sh4, institution, staff_rows, month, year)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def _staff_rows(self, month: int, year: int, import_id: int | None = None) -> list[dict[str, Any]]:
        """Todo el personal activo. Sin marcaciones del mes = celdas vacias."""
        attendance_rows = attendance_service.list_month(month, year, import_id=import_id)
        days_by_staff: dict[int, dict[str, dict[str, Any]]] = defaultdict(dict)
        for item in attendance_rows:
            days_by_staff[item["staff_member_id"]][item["attendance_date"]] = item
        rows = []
        for staff in staff_member_service.list(is_active="Y"):
            rows.append({**staff, "days": days_by_staff.get(staff["id"], {})})
        return rows

    def _header_annex03(
        self,
        sheet: Any,
        institution: dict[str, Any],
        month: int,
        year: int,
        end_column: int,
    ) -> int:
        """Cabecera Anexo 03 con columnas fijas pedidas:

        Labels A-B valores C-F  |  MES/LUGAR/DEP G-J valores K-P
        ANO/PROV Q-S valores T-Y  |  TURNO/DIS AA-AC valores AD-AH
        """
        ugel = _safe(institution.get("ugel"))
        school = _safe(institution.get("school_name"))
        level = _safe(institution.get("education_level"))
        modular = _safe(institution.get("modular_code"))
        address = _safe(institution.get("address") or institution.get("lugar"))
        department = _safe(institution.get("department") or "PUNO")
        province = _safe(institution.get("province"))
        district = _safe(institution.get("district"))
        shift = _safe(institution.get("shift_name"))

        # Asegurar ancho minimo para llegar a AH (col 34)
        end_column = max(end_column, 34)

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
        c = sheet.cell(1, 1, LEGAL_HEADER)
        c.font = Font(bold=True, size=7, name="Calibri")
        c.alignment = CENTER
        sheet.row_dimensions[1].height = 20

        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
        c = sheet.cell(2, 1, "ANEXO 03")
        c.font = Font(bold=True, size=10, name="Calibri")
        c.alignment = CENTER

        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_column)
        c = sheet.cell(3, 1, "FORMATO 01: REPORTE DE ASISTENCIA DETALLADO")
        c.fill = FILL_TITLE
        c.font = FONT_TITLE
        c.alignment = CENTER
        sheet.row_dimensions[3].height = 20

        def _lab(row: int, c1: int, c2: int, text: str) -> None:
            if c2 > c1:
                sheet.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            cell = sheet.cell(row, c1, text)
            cell.font = FONT_BLACK_BOLD
            cell.alignment = NOWRAP_LEFT

        def _val(row: int, c1: int, c2: int, text: str) -> None:
            if c2 > c1:
                sheet.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            cell = sheet.cell(row, c1, text)
            cell.font = FONT_RED_BOLD
            cell.alignment = NOWRAP_LEFT

        # --- Fila 4 ---
        # A-B UGEL: | C-F valor | G-J MES: | K-P valor | Q-S ANO: | T-Y valor | AA-AC TURNO: | AD-AH valor
        _lab(4, 1, 2, "UGEL:")
        _val(4, 3, 6, ugel)
        _lab(4, 7, 10, "MES:")
        _val(4, 11, 16, MONTH_NAMES[month - 1])
        _lab(4, 17, 19, "ANO:")
        _val(4, 20, 25, str(year))
        _lab(4, 27, 29, "TURNO:")
        _val(4, 30, 34, shift)

        # --- Fila 5 ---
        # A-B IE: | C-F valor | G-J LUGAR: | K-P valor
        _lab(5, 1, 2, "IE:")
        _val(5, 3, 6, school)
        _lab(5, 7, 10, "LUGAR:")
        _val(5, 11, 16, address)

        # --- Fila 6 ---
        # A-B NIVEL: | C-F valor | G-J DEP: | K-P valor | Q-S PROV: | T-Y valor | AA-AC DIS: | AD-AH valor
        _lab(6, 1, 2, "NIVEL:")
        _val(6, 3, 6, level)
        _lab(6, 7, 10, "DEP:")
        _val(6, 11, 16, department)
        _lab(6, 17, 19, "PROV:")
        _val(6, 20, 25, province)
        _lab(6, 27, 29, "DIS:")
        _val(6, 30, 34, district)

        # --- Fila 7 ---
        # A-B COD.MOD: | C-F valor
        _lab(7, 1, 2, "COD.MOD:")
        _val(7, 3, 6, modular)

        for r in range(4, 8):
            sheet.row_dimensions[r].height = 14
        sheet.row_dimensions[8].height = 4
        return 9

    def _header_annex04(
        self,
        sheet: Any,
        institution: dict[str, Any],
        month: int,
        year: int,
        end_column: int,
    ) -> int:
        """Cabecera Anexo 04 con columnas fijas:

        Labels A-B valores C-E  |  MES/LUGAR/DEP F-G valores H-I
        ANO/PROV K valores L-M  |  TURNO/DIS N valores O-P
        """
        ugel = _safe(institution.get("ugel"))
        school = _safe(institution.get("school_name"))
        level = _safe(institution.get("education_level"))
        modular = _safe(institution.get("modular_code"))
        address = _safe(institution.get("address") or institution.get("lugar"))
        department = _safe(institution.get("department") or "PUNO")
        province = _safe(institution.get("province"))
        district = _safe(institution.get("district"))
        shift = _safe(institution.get("shift_name"))

        end_column = max(end_column, 16)

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
        c = sheet.cell(1, 1, LEGAL_HEADER)
        c.font = Font(bold=True, size=7, name="Calibri")
        c.alignment = CENTER
        sheet.row_dimensions[1].height = 20

        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
        c = sheet.cell(2, 1, "ANEXO 04")
        c.font = Font(bold=True, size=10, name="Calibri")
        c.alignment = CENTER

        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_column)
        c = sheet.cell(
            3, 1,
            "FORMATO 02: REPORTE CONSOLIDADO DE INASISTENCIAS, TARDANZAS Y PERMISOS SIN GOCE DE REMUNERACION",
        )
        c.fill = FILL_TITLE
        c.font = FONT_TITLE
        c.alignment = CENTER
        sheet.row_dimensions[3].height = 20

        def _lab(row: int, c1: int, c2: int, text: str) -> None:
            if c2 > c1:
                sheet.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            cell = sheet.cell(row, c1, text)
            cell.font = FONT_BLACK_BOLD
            cell.alignment = NOWRAP_LEFT

        def _val(row: int, c1: int, c2: int, text: str) -> None:
            if c2 > c1:
                sheet.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            cell = sheet.cell(row, c1, text)
            cell.font = FONT_RED_BOLD
            cell.alignment = NOWRAP_LEFT

        # --- Fila 4 ---
        # A-B UGEL: | C-E valor | F-G MES: | H-I valor | K ANO: | L-M valor | N TURNO: | O-P valor
        _lab(4, 1, 2, "UGEL:")
        _val(4, 3, 5, ugel)
        _lab(4, 6, 7, "MES:")
        _val(4, 8, 9, MONTH_NAMES[month - 1])
        _lab(4, 11, 11, "ANO:")
        _val(4, 12, 13, str(year))
        _lab(4, 14, 14, "TURNO:")
        _val(4, 15, 16, shift)

        # --- Fila 5 ---
        # A-B IE: | C-E valor | F-G LUGAR: | H-I valor
        _lab(5, 1, 2, "IE:")
        _val(5, 3, 5, school)
        _lab(5, 6, 7, "LUGAR:")
        _val(5, 8, 9, address)

        # --- Fila 6 ---
        # A-B NIVEL: | C-E valor | F-G DEP: | H-I valor | K PROV: | L-M valor | N DIS: | O-P valor
        _lab(6, 1, 2, "NIVEL:")
        _val(6, 3, 5, level)
        _lab(6, 6, 7, "DEP:")
        _val(6, 8, 9, department)
        _lab(6, 11, 11, "PROV:")
        _val(6, 12, 13, province)
        _lab(6, 14, 14, "DIS:")
        _val(6, 15, 16, district)

        # --- Fila 7 ---
        # A-B COD.MOD: | C-E valor
        _lab(7, 1, 2, "COD.MOD:")
        _val(7, 3, 5, modular)

        for r in range(4, 8):
            sheet.row_dimensions[r].height = 14
        sheet.row_dimensions[8].height = 4
        return 9

    def _write_attendance_sheet(
        self,
        sheet: Any,
        institution: dict[str, Any],
        staff_rows: list[dict[str, Any]],
        month: int,
        year: int,
    ) -> None:
        days = monthrange(year, month)[1]
        first_day = 7
        end_col = first_day + days - 1
        table_start = self._header_annex03(sheet, institution, month, year, end_col)

        fixed = ["N°", "DNI", "APELLIDOS Y NOMBRES", "CARGO", "CONDICION\nLABORAL", "JORNADA\nLABORAL"]
        hr = table_start
        dr = table_start + 1
        wr = table_start + 2

        for col, h in enumerate(fixed, 1):
            sheet.merge_cells(start_row=hr, start_column=col, end_row=dr, end_column=col)
            cell = sheet.cell(hr, col, h)
            cell.fill = FILL_HEADER
            cell.font = FONT_BLACK_BOLD
            cell.alignment = CENTER
            cell.border = BORDER_THIN
            sheet.cell(dr, col).border = BORDER_THIN
            sheet.cell(dr, col).fill = FILL_HEADER

        sheet.merge_cells(start_row=hr, start_column=first_day, end_row=hr, end_column=end_col)
        cell = sheet.cell(hr, first_day, "DIAS CALENDARIO")
        cell.fill = FILL_HEADER
        cell.font = FONT_BLACK_BOLD
        cell.alignment = CENTER
        for c in range(first_day, end_col + 1):
            sheet.cell(hr, c).border = BORDER_THIN
            sheet.cell(hr, c).fill = FILL_HEADER

        wd = ["lu.", "ma.", "mi.", "ju.", "vi.", "sa.", "do."]
        for day in range(1, days + 1):
            col = first_day + day - 1
            cell = sheet.cell(dr, col, day)
            cell.fill = FILL_HEADER
            cell.font = FONT_BLACK_BOLD
            cell.alignment = CENTER
            cell.border = BORDER_THIN
            w_idx = date(year, month, day).weekday()
            is_weekend = w_idx in (5, 6)
            wcell = sheet.cell(wr, col, wd[w_idx])
            wcell.font = FONT_RED_WEEKEND if is_weekend else FONT_TINY
            wcell.alignment = CENTER
            wcell.border = BORDER_THIN
            wcell.fill = FILL_WEEKEND if is_weekend else FILL_SUB

        sheet.row_dimensions[hr].height = 16
        sheet.row_dimensions[dr].height = 12
        sheet.row_dimensions[wr].height = 12

        data0 = wr + 1
        codes = {
            "no_record": "-", "present": "A", "late": "T", "absent": "I",
            "justified": "J", "leave": "J", "unpaid_leave": "LS",
            "permission": "P", "strike": "H", "holiday": "F",
        }
        for i, staff in enumerate(staff_rows, 1):
            row = data0 + i - 1
            status = _safe(staff.get("employment_status"))
            if len(status) > 15:
                status = status.split()[0]
            vals = [
                i,
                staff.get("dni") or "",
                _safe(self._full_name(staff)),
                _safe(staff.get("job_title")),
                status,
                "",
            ]
            for col, v in enumerate(vals, 1):
                cell = sheet.cell(row, col, v)
                cell.border = BORDER_THIN
                cell.font = FONT_NORMAL
                cell.alignment = CENTER if col in (1, 2, 5, 6) else LEFT
            for day in range(1, days + 1):
                att = staff["days"].get(date(year, month, day).isoformat())
                st = att.get("status") if att else ""
                cell = sheet.cell(row, first_day + day - 1, codes.get(st, ""))
                cell.alignment = CENTER
                cell.border = BORDER_THIN
                cell.font = FONT_TINY
            sheet.row_dimensions[row].height = 14

        last = data0 + max(len(staff_rows), 1) - 1
        widths = [4, 10, 26, 12, 12, 9]
        for col, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = w
        for d in range(days):
            sheet.column_dimensions[get_column_letter(first_day + d)].width = 2.8

        self._print_a4(sheet, hr, last, end_col)
        sheet.freeze_panes = f"{get_column_letter(first_day)}{data0}"

    def _write_consolidated_sheet(
        self,
        sheet: Any,
        institution: dict[str, Any],
        staff_rows: list[dict[str, Any]],
        month: int,
        year: int,
    ) -> None:
        end_col = 16
        table_start = self._header_annex04(sheet, institution, month, year, end_col)
        hr1 = table_start
        hr2 = table_start + 1

        single_headers = [
            (1, "N°"), (2, "DNI"), (3, "APELLIDOS Y NOMBRES"),
            (4, "CARGO"), (5, "CONDICION\nLABORAL"), (6, "JORNADA\nLABORAL"),
        ]
        for col, h in single_headers:
            sheet.merge_cells(start_row=hr1, start_column=col, end_row=hr2, end_column=col)
            c = sheet.cell(hr1, col, h)
            c.fill = FILL_HEADER
            c.font = FONT_BLACK_BOLD
            c.alignment = CENTER
            c.border = BORDER_THIN
            sheet.cell(hr2, col).border = BORDER_THIN
            sheet.cell(hr2, col).fill = FILL_HEADER

        groups = [
            (7, 7, "INASISTENCIAS\nJUSTIFICADAS", ["DIAS"]),
            (8, 10, "LICENCIAS", ["CON\nGOCE", "SIN GOCE", "DU"]),
            (11, 11, "FALTAS", ["DIAS"]),
            (12, 12, "TARDANZAS", ["MINUTOS (*)"]),
            (13, 14, "PERMISOS SG", ["HORAS (*)", "MINUTOS (*)"]),
            (15, 15, "HUELGA PARO", ["DIAS"]),
        ]
        for start_c, end_c, title, subs in groups:
            sheet.merge_cells(start_row=hr1, start_column=start_c, end_row=hr1, end_column=end_c)
            c = sheet.cell(hr1, start_c, title)
            c.fill = FILL_HEADER
            c.font = FONT_BLACK_BOLD
            c.alignment = CENTER
            for col_idx in range(start_c, end_c + 1):
                sheet.cell(hr1, col_idx).border = BORDER_THIN
                sheet.cell(hr1, col_idx).fill = FILL_HEADER
            for idx, sub_title in enumerate(subs):
                sc = sheet.cell(hr2, start_c + idx, sub_title)
                sc.fill = FILL_HEADER
                sc.font = FONT_BLACK_BOLD
                sc.alignment = CENTER
                sc.border = BORDER_THIN

        sheet.merge_cells(start_row=hr1, start_column=16, end_row=hr2, end_column=16)
        c = sheet.cell(hr1, 16, "Observaciones")
        c.fill = FILL_HEADER
        c.font = FONT_BLACK_BOLD
        c.alignment = CENTER
        c.border = BORDER_THIN
        sheet.cell(hr2, 16).border = BORDER_THIN
        sheet.cell(hr2, 16).fill = FILL_HEADER

        sheet.row_dimensions[hr1].height = 20
        sheet.row_dimensions[hr2].height = 22

        data0 = hr2 + 1
        for i, staff in enumerate(staff_rows, 1):
            counts = Counter(x["status"] for x in staff["days"].values())
            late_min = sum(
                x.get("late_minutes", 0)
                for x in staff["days"].values()
                if x.get("status") == "late"
            )
            status = _safe(staff.get("employment_status"))
            if len(status) > 15:
                status = status.split()[0]
            row = data0 + i - 1
            vals = [
                i,
                staff.get("dni") or "",
                _safe(self._full_name(staff)),
                _safe(staff.get("job_title")),
                status,
                "",
                counts.get("justified", 0),
                counts.get("leave", 0),
                counts.get("unpaid_leave", 0), 0,
                counts.get("absent", 0),
                late_min,
                0, 0, counts.get("strike", 0), "",
            ]
            for col, v in enumerate(vals, 1):
                cell = sheet.cell(row, col, v)
                cell.border = BORDER_THIN
                cell.font = FONT_NORMAL
                cell.alignment = CENTER if col in (1, 2, 5, 6) or col >= 7 else LEFT
            sheet.row_dimensions[row].height = 14

        last = data0 + max(len(staff_rows), 1) - 1
        widths = [4, 10, 26, 12, 12, 9, 10, 7, 8, 5, 7, 10, 9, 10, 8, 18]
        for col, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = w

        self._print_a4(sheet, hr1, last, end_col)
        sheet.freeze_panes = f"A{data0}"

    def _print_a4(self, sheet: Any, start_row: int, end_row: int, end_col: int) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins.left = 0.2
        sheet.page_margins.right = 0.2
        sheet.page_margins.top = 0.3
        sheet.page_margins.bottom = 0.3
        sheet.print_title_rows = f"1:{start_row}"
        sheet.print_area = f"A1:{get_column_letter(end_col)}{end_row}"

    def _staff_member(self, staff_member_id: int) -> dict[str, Any]:
        try:
            return staff_member_service.get(staff_member_id)
        except StaffMemberNotFoundError:
            return {"dni": None, "last_names": "Unknown", "first_names": "Staff"}

    def _full_name(self, staff_member: dict[str, Any]) -> str:
        last = staff_member.get("last_names") or ""
        first = staff_member.get("first_names") or ""
        return f"{last}, {first}".strip(", ")

    def _institution(self) -> dict[str, Any]:
        try:
            inst = institution_repository.get_active()
            if inst:
                return {**DEMO_INSTITUTION, **inst}
            return DEMO_INSTITUTION
        except OracleRepositoryError:
            return DEMO_INSTITUTION


report_service = ReportService()
